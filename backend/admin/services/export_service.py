"""
Export Service - Generate deterministic exports of admitted articles
Handles CSV and XLSX exports with time-range filtering
"""

import csv
import json
from io import BytesIO, StringIO
from datetime import datetime, timedelta, date
from typing import Dict, List, Optional, Tuple, Any

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False


class ExportService:
    """Export admitted articles in CSV/XLSX formats with time-range filtering"""

    # Column definitions: (field_name, db_column, display_name, transform_function)
    EXPORT_COLUMNS = [
        ('id', 'id', 'ID', None),
        ('published_at', 'date_published', 'Published Date', 'format_date'),
        ('source_name', 'source_name', 'Source', None),
        ('url', 'url', 'URL', None),
        ('title', 'title', 'Title', None),
        ('category', 'category', 'Category', None),
        ('state_codes', 'state_codes', 'States', 'flatten_states'),
        ('sector', 'sector', 'Sector', None),
        ('importance_score', 'importance_score', 'Importance Score', 'format_score'),
        ('premium_processed', 'premium_processed', 'Premium Processed', 'format_bool'),
        ('summary', 'summary', 'Summary', None),
    ]

    def __init__(self, db_session):
        """
        Initialize export service

        Args:
            db_session: SQLAlchemy database session
        """
        self.db = db_session

    def calculate_date_range(self, days: Optional[int] = None,
                            start_date: Optional[str] = None,
                            end_date: Optional[str] = None) -> Tuple[date, date]:
        """
        Calculate date range for filtering

        Args:
            days: Number of days to look back from today
            start_date: Start date string (YYYY-MM-DD)
            end_date: End date string (YYYY-MM-DD)

        Returns:
            Tuple of (start_date, end_date) as date objects

        Raises:
            ValueError: If date parameters are invalid or conflicting
        """
        # Validate mutually exclusive parameters
        if days and (start_date or end_date):
            raise ValueError("Cannot specify both 'days' and 'start_date/end_date'")

        if days:
            # Calculate range from days parameter
            if days not in [7, 14, 21, 30]:
                raise ValueError(f"Invalid days value: {days}. Must be 7, 14, 21, or 30")
            end = datetime.utcnow().date()
            start = end - timedelta(days=days)
            return start, end

        if start_date and end_date:
            # Parse custom date range
            try:
                start = datetime.strptime(start_date, '%Y-%m-%d').date()
                end = datetime.strptime(end_date, '%Y-%m-%d').date()
            except ValueError as e:
                raise ValueError(f"Invalid date format. Use YYYY-MM-DD: {e}")

            if start > end:
                raise ValueError("start_date must be before or equal to end_date")

            return start, end

        raise ValueError("Must specify either 'days' or both 'start_date' and 'end_date'")

    def get_articles(self, start_date: date, end_date: date,
                    approved_only: bool = True) -> List[Any]:
        """
        Fetch articles within date range

        Args:
            start_date: Filter articles published on or after this date
            end_date: Filter articles published on or before this date
            approved_only: If True, only return approved articles (default: True)

        Returns:
            List of Update model instances
        """
        # Import here to avoid circular imports
        from app import Update

        # Base query: exclude soft-deleted, filter by date range
        query = Update.query.filter(
            (Update.is_deleted == False) | (Update.is_deleted == None),
            Update.date_published.isnot(None),
            Update.date_published >= start_date,
            Update.date_published <= end_date
        )

        # Optional: filter by approval status
        if approved_only:
            query = query.filter(Update.is_approved == True)

        # Order by date_published ascending for deterministic output
        articles = query.order_by(Update.date_published.asc(), Update.id.asc()).all()

        return articles

    def transform_value(self, article: Any, db_column: str,
                       transform_fn: Optional[str]) -> str:
        """
        Transform database value to export format

        Args:
            article: Update model instance
            db_column: Database column name
            transform_fn: Transformation function name

        Returns:
            Transformed string value
        """
        # Get value from database
        value = getattr(article, db_column, None)

        # Apply transformations
        if transform_fn == 'format_date':
            return value.isoformat() if value else ''
        elif transform_fn == 'flatten_states':
            if value:
                try:
                    states = json.loads(value)
                    return ', '.join(states) if isinstance(states, list) else ''
                except (json.JSONDecodeError, TypeError):
                    return ''
            return ''
        elif transform_fn == 'format_score':
            if value is not None:
                return str(round(float(value)))
            return '0'
        elif transform_fn == 'format_bool':
            return 'Yes' if value else 'No'
        else:
            # Default: convert to string
            return str(value) if value is not None else ''

    def export_csv(self, articles: List[Any]) -> BytesIO:
        """
        Generate CSV export

        Args:
            articles: List of Update model instances

        Returns:
            BytesIO buffer containing CSV data
        """
        # Use StringIO first, then convert to BytesIO for consistent API
        string_buffer = StringIO()
        writer = csv.writer(string_buffer, quoting=csv.QUOTE_MINIMAL)

        # Write header row
        headers = [col[2] for col in self.EXPORT_COLUMNS]  # display_name
        writer.writerow(headers)

        # Write data rows
        for article in articles:
            row = []
            for field_name, db_column, display_name, transform_fn in self.EXPORT_COLUMNS:
                value = self.transform_value(article, db_column, transform_fn)
                row.append(value)
            writer.writerow(row)

        # Convert to BytesIO
        byte_buffer = BytesIO()
        byte_buffer.write(string_buffer.getvalue().encode('utf-8-sig'))  # UTF-8 BOM for Excel
        byte_buffer.seek(0)

        return byte_buffer

    def export_xlsx(self, articles: List[Any]) -> BytesIO:
        """
        Generate XLSX export with formatting

        Args:
            articles: List of Update model instances

        Returns:
            BytesIO buffer containing XLSX data

        Raises:
            RuntimeError: If openpyxl is not available
        """
        if not XLSX_AVAILABLE:
            raise RuntimeError("openpyxl library not available. Install with: pip install openpyxl")

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Articles"

        # Style for header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Write header row with styling
        headers = [col[2] for col in self.EXPORT_COLUMNS]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill

        # Write data rows
        for row_idx, article in enumerate(articles, start=2):
            for col_idx, (field_name, db_column, display_name, transform_fn) in enumerate(self.EXPORT_COLUMNS, start=1):
                value = self.transform_value(article, db_column, transform_fn)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for col_idx, (field_name, db_column, display_name, transform_fn) in enumerate(self.EXPORT_COLUMNS, start=1):
            if field_name in ['summary', 'title', 'url']:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 50
            else:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15

        # Save to BytesIO
        byte_buffer = BytesIO()
        wb.save(byte_buffer)
        byte_buffer.seek(0)

        return byte_buffer

    def generate_export(self, format: str, days: Optional[int] = None,
                       start_date: Optional[str] = None, end_date: Optional[str] = None,
                       approved_only: bool = True) -> Tuple[BytesIO, Dict[str, Any]]:
        """
        Main export generation method

        Args:
            format: 'csv' or 'xlsx'
            days: Number of days to look back
            start_date: Custom start date (YYYY-MM-DD)
            end_date: Custom end date (YYYY-MM-DD)
            approved_only: Filter to approved articles only

        Returns:
            Tuple of (file_buffer, stats_dict)

        Raises:
            ValueError: If parameters are invalid
            RuntimeError: If XLSX export requested but openpyxl unavailable
        """
        # Validate format
        if format not in ['csv', 'xlsx']:
            raise ValueError(f"Invalid format: {format}. Must be 'csv' or 'xlsx'")

        if format == 'xlsx' and not XLSX_AVAILABLE:
            raise RuntimeError("XLSX export requires openpyxl library")

        # Calculate date range
        date_start, date_end = self.calculate_date_range(days, start_date, end_date)

        # Fetch articles
        articles = self.get_articles(date_start, date_end, approved_only)

        # Generate export
        if format == 'csv':
            file_buffer = self.export_csv(articles)
        else:
            file_buffer = self.export_xlsx(articles)

        # Compile stats
        stats = {
            'format': format,
            'start_date': date_start.isoformat(),
            'end_date': date_end.isoformat(),
            'approved_only': approved_only,
            'article_count': len(articles),
            'file_size_bytes': file_buffer.getbuffer().nbytes
        }

        return file_buffer, stats

    # ==================== JSON-based Export Methods ====================

    @staticmethod
    def generate_id_from_url(url: str) -> str:
        """Generate numeric ID from URL hash"""
        return str(abs(hash(url)) % (10 ** 8))  # 8-digit numeric ID

    @staticmethod
    def transform_value_from_dict(article: Dict, db_column: str,
                                 transform_fn: Optional[str]) -> str:
        """
        Transform dict value to export format

        Args:
            article: Article dict from JSON
            db_column: Field name in dict
            transform_fn: Transformation function name

        Returns:
            Transformed string value
        """
        value = article.get(db_column)

        if transform_fn == 'format_date':
            if value:
                # Already in YYYY-MM-DD format from JSON
                return value
            return ''
        elif transform_fn == 'flatten_states':
            if value and isinstance(value, list):
                return ', '.join(value)
            return ''
        elif transform_fn == 'format_score':
            if value is not None:
                return str(round(float(value)))
            return '0'
        elif transform_fn == 'format_bool':
            return 'Yes' if value else 'No'
        else:
            # Default: convert to string
            return str(value) if value is not None else ''

    @staticmethod
    def export_csv_from_dicts(articles: List[Dict]) -> BytesIO:
        """
        Generate CSV export from article dicts

        Args:
            articles: List of article dicts from JSON

        Returns:
            BytesIO buffer containing CSV data
        """
        string_buffer = StringIO()
        writer = csv.writer(string_buffer, quoting=csv.QUOTE_MINIMAL)

        # Write header row
        headers = [col[2] for col in ExportService.EXPORT_COLUMNS]  # display_name
        writer.writerow(headers)

        # Write data rows
        for article in articles:
            row = []
            for field_name, db_column, display_name, transform_fn in ExportService.EXPORT_COLUMNS:
                if db_column == 'id':
                    # Generate ID from URL
                    value = ExportService.generate_id_from_url(article.get('url', ''))
                else:
                    value = ExportService.transform_value_from_dict(article, db_column, transform_fn)
                row.append(value)
            writer.writerow(row)

        # Convert to BytesIO
        byte_buffer = BytesIO()
        byte_buffer.write(string_buffer.getvalue().encode('utf-8-sig'))  # UTF-8 BOM for Excel
        byte_buffer.seek(0)

        return byte_buffer

    @staticmethod
    def export_xlsx_from_dicts(articles: List[Dict]) -> BytesIO:
        """
        Generate XLSX export from article dicts

        Args:
            articles: List of article dicts from JSON

        Returns:
            BytesIO buffer containing XLSX data

        Raises:
            RuntimeError: If openpyxl is not available
        """
        if not XLSX_AVAILABLE:
            raise RuntimeError("openpyxl library not available. Install with: pip install openpyxl")

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Articles"

        # Style for header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Write header row with styling
        headers = [col[2] for col in ExportService.EXPORT_COLUMNS]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill

        # Write data rows
        for row_idx, article in enumerate(articles, start=2):
            for col_idx, (field_name, db_column, display_name, transform_fn) in enumerate(ExportService.EXPORT_COLUMNS, start=1):
                if db_column == 'id':
                    # Generate ID from URL
                    value = ExportService.generate_id_from_url(article.get('url', ''))
                else:
                    value = ExportService.transform_value_from_dict(article, db_column, transform_fn)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for col_idx, (field_name, db_column, display_name, transform_fn) in enumerate(ExportService.EXPORT_COLUMNS, start=1):
            if field_name in ['summary', 'title', 'url']:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 50
            else:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15

        # Save to BytesIO
        byte_buffer = BytesIO()
        wb.save(byte_buffer)
        byte_buffer.seek(0)

        return byte_buffer

    @staticmethod
    def generate_export_from_json(data_manager, format: str, days: Optional[int] = None,
                                  start_date: Optional[str] = None, end_date: Optional[str] = None,
                                  approved_only: bool = True, category: Optional[str] = None,
                                  state: Optional[str] = None) -> Tuple[BytesIO, Dict[str, Any]]:
        """
        Generate export from JSON files via DataManager

        Args:
            data_manager: DataManager instance
            format: 'csv' or 'xlsx'
            days: Number of days to look back
            start_date: Custom start date (YYYY-MM-DD)
            end_date: Custom end date (YYYY-MM-DD)
            approved_only: Filter to approved articles only
            category: Optional category filter
            state: Optional state filter

        Returns:
            Tuple of (file_buffer, stats_dict)

        Raises:
            ValueError: If parameters are invalid
            RuntimeError: If XLSX export requested but openpyxl unavailable
        """
        # Validate format
        if format not in ['csv', 'xlsx']:
            raise ValueError(f"Invalid format: {format}. Must be 'csv' or 'xlsx'")

        if format == 'xlsx' and not XLSX_AVAILABLE:
            raise RuntimeError("XLSX export requires openpyxl library")

        # Calculate date range
        if days and (start_date or end_date):
            raise ValueError("Cannot specify both 'days' and 'start_date/end_date'")

        if days:
            if days not in [7, 14, 21, 30]:
                raise ValueError(f"Invalid days value: {days}. Must be 7, 14, 21, or 30")
            date_end = datetime.utcnow().date()
            date_start = date_end - timedelta(days=days)
        elif start_date and end_date:
            try:
                date_start = datetime.strptime(start_date, '%Y-%m-%d').date()
                date_end = datetime.strptime(end_date, '%Y-%m-%d').date()
            except ValueError as e:
                raise ValueError(f"Invalid date format. Use YYYY-MM-DD: {e}")

            if date_start > date_end:
                raise ValueError("start_date must be before or equal to end_date")
        else:
            raise ValueError("Must specify either 'days' or both 'start_date' and 'end_date'")

        # Get articles from DataManager with optional filters
        filters = {
            'category': category,
            'state': state,
            'search': None
        }
        all_articles = data_manager.get_all_articles(filters)

        # Filter by date range
        filtered = []
        for article in all_articles:
            # Get date from article
            date_str = article.get('date_published') or article.get('date')
            if date_str:
                try:
                    article_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                    if date_start <= article_date <= date_end:
                        filtered.append(article)
                except (ValueError, TypeError):
                    # Skip articles with invalid dates
                    continue

        # Sort by date for deterministic output
        filtered.sort(key=lambda x: (
            x.get('date_published') or x.get('date') or '9999-99-99',
            x.get('url', '')
        ))

        # Generate export
        if format == 'csv':
            file_buffer = ExportService.export_csv_from_dicts(filtered)
        else:
            file_buffer = ExportService.export_xlsx_from_dicts(filtered)

        # Compile stats
        stats = {
            'format': format,
            'start_date': date_start.isoformat(),
            'end_date': date_end.isoformat(),
            'approved_only': approved_only,
            'article_count': len(filtered),
            'file_size_bytes': file_buffer.getbuffer().nbytes
        }

        return file_buffer, stats
